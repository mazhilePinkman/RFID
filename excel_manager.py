import os
import openpyxl
import tkinter as tk  # 添加这行
import tkinter.messagebox as messagebox
from datetime import datetime


class ExcelManager:
    """管理Excel文件操作的类"""

    def __init__(self, app):
        self.app = app
        self.current_excel_path = None

    def generate_excel(self):
        """生成Excel文件，根据盘点模式添加或更新记录，同一TID只占一行"""
        if not self.app.tree.get_children():
            messagebox.showwarning("没有数据", "没有扫描数据可供导出")
            return False

        try:
            filename = f"RFID_{datetime.now().strftime('%Y%m%d')}.xlsx"
            file_path = os.path.abspath(filename)
            mode = self.app.inventory_mode.get()

            # 读取或新建Excel
            if os.path.exists(file_path):
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                tid_row_map = {}
                max_id = 0
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if len(row) >= 2:
                        tid = str(row[1]).strip().upper()
                        tid_row_map[tid] = row_idx
                        if row[0] and (isinstance(row[0], int) or str(row[0]).isdigit()):
                            row_id = int(row[0])
                            if row_id > max_id:
                                max_id = row_id
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "RFID标签"
                headers = ["编号", "TID", "入场时间", "出场时间", "项目地址", "备注"]
                ws.append(headers)
                tid_row_map = {}
                max_id = 0

            updated_count = 0
            added_count = 0

            for tid, data in self.app.tid_counts.items():
                tid_upper = tid.upper()
                if tid_upper in tid_row_map:
                    row_idx = tid_row_map[tid_upper]
                    # 入库写入入场时间，出库写入出场时间
                    if mode == "入库":
                        ws.cell(row=row_idx, column=3, value=data["last_time"])
                    else:
                        ws.cell(row=row_idx, column=4, value=data["last_time"])
                    updated_count += 1
                else:
                    max_id += 1
                    if mode == "入库":
                        ws.append([
                            max_id,
                            tid,
                            data["last_time"],
                            "",
                            "",
                            ""
                        ])
                    else:
                        ws.append([
                            max_id,
                            tid,
                            "",
                            data["last_time"],
                            "",
                            ""
                        ])
                    added_count += 1

            wb.save(file_path)
            self.current_excel_path = file_path
            self.app.upload_btn.config(state=tk.NORMAL)
            self.app.log_message(
                f"已生成/更新{mode}Excel文件: {os.path.basename(file_path)}，"
                f"更新{updated_count}条，新增{added_count}条"
            )
            success_msg = (
                f"已成功生成/更新{mode}Excel文件！\n\n"
                f"文件名: {os.path.basename(file_path)}\n"
                f"保存位置: {os.path.dirname(file_path)}\n"
                f"本次写入/更新标签数: {updated_count + added_count}"
            )
            messagebox.showinfo("操作成功", success_msg)
            return True

        except Exception as e:
            error_msg = f"生成Excel时出错: {str(e)}"
            self.app.log_message(error_msg)
            messagebox.showerror("生成失败", error_msg)
            return False
